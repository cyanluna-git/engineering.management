"""
Seed resource plans from Excel template to remote production DB.

Reads: docs/IS Resource Plan Template_V1.2.xlsx
Target: Remote PostgreSQL on 10.182.252.32 via SSH

Steps:
  1. Add new project roles
  2. Update project lifecycle statuses (bulk + 6 specific)
  3. Update project start_month / end_month
  4. Upsert gate milestones
  5. Delete dummy resource_plans + histories
  6. Insert resource plans from Excel
"""

import subprocess
import sys
import os

# Add project root to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from datetime import datetime, date
import calendar
import json

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "docs", "IS Resource Plan Template_V1.2.xlsx"
)

SSH_CMD = "ssh atlasAdmin@10.182.252.32"
PSQL_CMD = 'docker exec -i edwards-postgres psql -U postgres edwards'

# ── Project Mapping ──────────────────────────────────────────────────────────
PROJECT_MAP = {
    "Havasu V00": {
        "id": "f9914dab-90fc-4e73-92f3-e17d638b9b02",
        "name": "Havasu",
        "status": "Complete",
        "start_month": "2025-06",
        "end_month": "2026-01",
    },
    "Tumalo Ph1 V00": {
        "id": "74a3027f-ac1f-4db6-b1df-7a3b699a1fb2",
        "name": "EUV Gen4 Phase 1 Tumalo",
        "status": "Active",
        "start_month": "2025-06",
        "end_month": "2026-08",
    },
    "Tumalo Ph2 V00": {
        "id": "dbf7bb73-6519-4e1b-a339-e7f666f526cf",
        "name": "EUV Gen4 Phase 2 Tumalo",
        "status": "Active",
        "start_month": "2025-07",
        "end_month": "2027-12",
    },
    "Kanara V00": {
        "id": "f480256a-a6b7-4c7e-bbf6-4ba6a5d4bf17",
        "name": "Gen4, Kanarra",
        "status": "Opportunity",
        "start_month": "2026-02",
        "end_month": "2028-06",
    },
    "SAVAS": {
        "id": "64055402-4035-44ed-bc58-df2a43c256b6",
        "name": "LPLN SAVAS",
        "status": "Lead",
        "start_month": "2026-04",
        "end_month": "2027-12",
    },
    "HRS": {
        "id": "c1380ec9-96dd-4575-afc4-41f7f5f83803",
        "name": "Hydrogen Recovery System",
        "status": "Active",
        "start_month": "2025-12",
        "end_month": "2027-12",
    },
}

# ── Gate Milestones (from Excel Gate Plan rows) ──────────────────────────────
GATE_DATA = {
    "f9914dab-90fc-4e73-92f3-e17d638b9b02": {  # Havasu
        "Gate 5": "2025-06",
        "Gate 6": "2025-12",
    },
    "74a3027f-ac1f-4db6-b1df-7a3b699a1fb2": {  # Tumalo Ph1
        "Gate 4": "2025-08",
        "Gate 5": "2026-02",
        "Gate 6": "2026-08",
    },
    "dbf7bb73-6519-4e1b-a339-e7f666f526cf": {  # Tumalo Ph2
        "Gate 2": "2025-07",
        "Gate 3": "2026-02",
        "Gate 4": "2026-12",
        "Gate 5": "2027-06",
        "Gate 6": "2027-12",
    },
    "f480256a-a6b7-4c7e-bbf6-4ba6a5d4bf17": {  # Kanara
        "Gate 1": "2026-02",
        "Gate 2": "2026-07",
        "Gate 3": "2026-12",
        "Gate 4": "2027-07",
        "Gate 5": "2027-12",
        "Gate 6": "2028-06",
    },
    "64055402-4035-44ed-bc58-df2a43c256b6": {  # SAVAS
        "Gate 1": "2026-04",
        "Gate 2": "2026-07",
        "Gate 3": "2026-12",
        "Gate 5": "2027-07",
        "Gate 6": "2027-12",
    },
    "c1380ec9-96dd-4575-afc4-41f7f5f83803": {  # HRS 7N (2nd Gate Plan row)
        "Gate 3": "2026-07",
        "Gate 4": "2027-01",
        "Gate 5": "2027-06",
        "Gate 6": "2027-12",
    },
}

# ── Excel Role → project_role_id ─────────────────────────────────────────────
ROLE_MAP = {
    "* Eng Manager": "PR_ENG_MGR",
    "* Manager": "PR_MGR",
    "PM": "PR_PM",
    "Tech Lead": "PR_TECH_LEAD",
    "Leader": "PR_TEAM_LEAD",
    "System engineer": "PR_SYS_ENG",
    "Software engineer": "PR_CTRL_ENG",
    "SW test engineer": "PR_SQA_ENG",
    "Service engineer": "PR_SVC_ENG",
    "Mechanical engineer": "PR_MECH_ENG",
    "Electrical engineer": "PR_HW_ENG",
    "Technician": "PR_TECHNICIAN",
}

# Names that are actually TBD positions (not real people)
TBD_NAMES = {
    "system engineer",
    "software engineer",
    "sw test engineer",
    "service engineer",
    "mechanical engineer",
    "electrical engineer",
    "technician",
    "project manager",
    "tech lead",
}

ADMIN_USER_ID = "f5503974-55a5-4fd6-91da-7916045fec86"


def decode_month(dt: datetime) -> tuple[int, int] | None:
    """Excel date trick: day encodes year (25->2025, 26->2026, 27->2027, 28->2028)"""
    if not hasattr(dt, "day"):
        return None
    year = 2000 + dt.day
    month = dt.month
    return (year, month)


def month_last_day(ym: str) -> str:
    """Convert 2025-06 to 2025-06-30"""
    y, m = int(ym[:4]), int(ym[5:7])
    last = calendar.monthrange(y, m)[1]
    return f"{y}-{m:02d}-{last:02d}"


def run_sql(sql: str) -> str:
    """Execute SQL on remote DB via SSH"""
    cmd = f'{SSH_CMD} "{PSQL_CMD}"'
    result = subprocess.run(
        cmd,
        shell=True,
        input=sql,
        capture_output=True,
        text=True,
        timeout=30,
    )
    if result.returncode != 0 and "ERROR" in result.stderr:
        print(f"SQL ERROR: {result.stderr}")
        print(f"SQL was: {sql[:500]}")
        sys.exit(1)
    return result.stdout


def escape_sql(s: str) -> str:
    """Escape single quotes for SQL"""
    if s is None:
        return "NULL"
    return s.replace("'", "''")


def build_sql() -> str:
    """Build the complete migration SQL"""
    sql_parts = []

    sql_parts.append("BEGIN;")
    sql_parts.append("")

    # ── Step 1: Add new project roles ────────────────────────────────────
    sql_parts.append("-- Step 1: Add new project roles")
    new_roles = [
        ("PR_ENG_MGR", "Engineering Manager", "Management"),
        ("PR_MGR", "Manager", "Management"),
        ("PR_TEAM_LEAD", "Team Leader", "Management"),
        ("PR_SVC_ENG", "Service Engineer", "Engineering"),
        ("PR_SQA_ENG", "SQA Engineer", "Engineering"),
        ("PR_TECHNICIAN", "Technician", "Support"),
    ]
    for role_id, name, category in new_roles:
        sql_parts.append(
            f"INSERT INTO project_roles (id, name, category, std_hourly_rate, is_active) "
            f"VALUES ('{role_id}', '{name}', '{category}', 0.0, true) "
            f"ON CONFLICT (id) DO NOTHING;"
        )
    sql_parts.append("")

    # ── Step 2: Update project lifecycle statuses ────────────────────────
    sql_parts.append("-- Step 2: Bulk status migration")
    sql_parts.append("UPDATE projects SET status = 'Active' WHERE status = 'InProgress';")
    sql_parts.append("UPDATE projects SET status = 'Complete' WHERE status = 'Completed';")
    sql_parts.append("UPDATE projects SET status = 'Lead' WHERE status = 'Prospective';")
    sql_parts.append("UPDATE projects SET status = 'Planning' WHERE status = 'Planned';")
    sql_parts.append("")

    sql_parts.append("-- Step 2b: Specific status for 6 target projects")
    for sheet_name, info in PROJECT_MAP.items():
        sql_parts.append(
            f"UPDATE projects SET status = '{info['status']}' WHERE id = '{info['id']}';"
        )
    sql_parts.append("")

    # ── Step 3: Update start_month / end_month ───────────────────────────
    sql_parts.append("-- Step 3: Update project start_month / end_month")
    for sheet_name, info in PROJECT_MAP.items():
        sql_parts.append(
            f"UPDATE projects SET start_month = '{info['start_month']}', "
            f"end_month = '{info['end_month']}' WHERE id = '{info['id']}';"
        )
    sql_parts.append("")

    # ── Step 4: Gate milestones ──────────────────────────────────────────
    sql_parts.append("-- Step 4: Delete existing milestones for target projects, then insert")
    project_ids = ", ".join(f"'{pid}'" for pid in GATE_DATA.keys())
    sql_parts.append(
        f"DELETE FROM project_milestones WHERE project_id IN ({project_ids});"
    )

    today = date.today()
    for project_id, gates in GATE_DATA.items():
        for gate_name, ym in gates.items():
            target_date = month_last_day(ym)
            is_key = gate_name in ("Gate 5", "Gate 6")
            gate_date = date.fromisoformat(target_date)
            status = "Completed" if gate_date < today else "Pending"
            sql_parts.append(
                f"INSERT INTO project_milestones "
                f"(project_id, name, type, target_date, status, is_key_gate) VALUES "
                f"('{project_id}', '{gate_name}', 'STD_GATE', "
                f"'{target_date}', '{status}', {str(is_key).lower()});"
            )
    sql_parts.append("")

    # ── Step 5: Delete all dummy resource plans ──────────────────────────
    sql_parts.append("-- Step 5: Delete all dummy resource plans and histories")
    sql_parts.append("DELETE FROM resource_plan_histories;")
    sql_parts.append("DELETE FROM resource_plans;")
    sql_parts.append("")

    # ── Step 6: Read Excel and insert resource plans ─────────────────────
    sql_parts.append("-- Step 6: Insert resource plans from Excel")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Get user lookup from DB
    user_result = run_sql(
        "SELECT id, name, position_id FROM users WHERE is_active = true;"
    )
    user_lookup = {}  # name_lower -> (id, position_id)
    for line in user_result.strip().split("\n"):
        if "|" not in line or "---" in line or "name" in line.lower()[:10]:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 3:
            uid, uname, pos_id = parts[0], parts[1], parts[2]
            user_lookup[uname.strip().lower()] = (uid, pos_id if pos_id else "JP_ENGINEER")

    total_inserted = 0

    for sheet_name, info in PROJECT_MAP.items():
        ws = wb[sheet_name]
        project_id = info["id"]

        # Build month columns
        months = {}
        for c in range(11, min(ws.max_column + 1, 62)):
            v = ws.cell(2, c).value
            ym = decode_month(v) if v else None
            if ym:
                months[c] = ym  # (year, month)

        # Process data rows (skip header row 2)
        sheet_count = 0
        for r in range(3, ws.max_row + 1):
            name_cell = ws.cell(r, 8).value
            if not name_cell:
                continue
            name = str(name_cell).strip()

            # Skip Gate Plan rows
            if "gate" in name.lower().replace(" ", ""):
                continue

            role_cell = ws.cell(r, 7).value
            role = str(role_cell).strip() if role_cell else ""

            project_role_id = ROLE_MAP.get(role)
            if not project_role_id:
                print(f"  WARNING: Unknown role '{role}' for '{name}' in {sheet_name}, skipping")
                continue

            # Determine user_id
            is_tbd = name.lower().strip() in TBD_NAMES
            user_id = "NULL"
            position_id = "JP_ENGINEER"

            if not is_tbd:
                # Try exact match
                lookup_key = name.lower().strip()
                if lookup_key in user_lookup:
                    uid, pos = user_lookup[lookup_key]
                    user_id = f"'{uid}'"
                    position_id = pos
                else:
                    # Try without trailing spaces, case variations
                    found = False
                    for db_name, (uid, pos) in user_lookup.items():
                        if db_name.replace(".", "").replace(" ", "") == lookup_key.replace(".", "").replace(" ", ""):
                            user_id = f"'{uid}'"
                            position_id = pos
                            found = True
                            break
                    if not found:
                        print(f"  WARNING: User '{name}' not found in DB for {sheet_name}, treating as TBD")
                        is_tbd = True

            # Process monthly FTE values
            for col, (year, month) in months.items():
                fte = ws.cell(r, col).value
                if fte is None or not isinstance(fte, (int, float)) or fte <= 0:
                    continue

                planned_hours = round(fte * 160.0, 1)

                sql_parts.append(
                    f"INSERT INTO resource_plans "
                    f"(project_id, year, month, position_id, project_role_id, "
                    f"user_id, planned_hours, created_by) VALUES "
                    f"('{project_id}', {year}, {month}, '{position_id}', "
                    f"'{project_role_id}', {user_id}, {planned_hours}, "
                    f"'{ADMIN_USER_ID}');"
                )
                sheet_count += 1

        total_inserted += sheet_count
        print(f"  {sheet_name}: {sheet_count} resource plan entries")

    sql_parts.append("")
    sql_parts.append("COMMIT;")

    print(f"\nTotal resource plan entries: {total_inserted}")
    return "\n".join(sql_parts)


def main() -> None:
    print("=" * 60)
    print("Excel Resource Plan Migration")
    print("=" * 60)
    print()

    print("Building SQL from Excel...")
    sql = build_sql()

    # Write SQL to file for review
    sql_file = os.path.join(os.path.dirname(__file__), "migration_resource_plans.sql")
    with open(sql_file, "w") as f:
        f.write(sql)
    print(f"\nSQL written to: {sql_file}")
    print(f"SQL size: {len(sql)} bytes, {sql.count(chr(10))} lines")

    # Execute
    print("\nExecuting on remote DB...")
    output = run_sql(sql)
    print(output)

    # Verify
    print("\n--- Verification ---")
    verify_sql = """
SELECT 'resource_plans' as tbl, COUNT(*) as cnt FROM resource_plans
UNION ALL
SELECT 'project_milestones', COUNT(*) FROM project_milestones
WHERE project_id IN ('f9914dab-90fc-4e73-92f3-e17d638b9b02','74a3027f-ac1f-4db6-b1df-7a3b699a1fb2','dbf7bb73-6519-4e1b-a339-e7f666f526cf','f480256a-a6b7-4c7e-bbf6-4ba6a5d4bf17','64055402-4035-44ed-bc58-df2a43c256b6','c1380ec9-96dd-4575-afc4-41f7f5f83803');

SELECT p.name, p.status, p.start_month, p.end_month,
       (SELECT COUNT(*) FROM resource_plans rp WHERE rp.project_id = p.id) as plan_count
FROM projects p
WHERE p.id IN ('f9914dab-90fc-4e73-92f3-e17d638b9b02','74a3027f-ac1f-4db6-b1df-7a3b699a1fb2','dbf7bb73-6519-4e1b-a339-e7f666f526cf','f480256a-a6b7-4c7e-bbf6-4ba6a5d4bf17','64055402-4035-44ed-bc58-df2a43c256b6','c1380ec9-96dd-4575-afc4-41f7f5f83803')
ORDER BY p.name;

SELECT name, id FROM project_roles WHERE id LIKE 'PR_%' ORDER BY name;

SELECT DISTINCT status, COUNT(*) FROM projects GROUP BY status ORDER BY status;
"""
    print(run_sql(verify_sql))


if __name__ == "__main__":
    main()
